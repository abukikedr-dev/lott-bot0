"""
Lottery Logbook Bot — v4.2 (Native Batching + Exports Restored)
"""

import os
import re
import time
import json
import logging
import itertools
import threading
import csv
import io
from typing import Optional, Dict, List
from threading import Thread

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import telebot
from telebot import types
import requests
import google.generativeai as genai

# ── Logging Setup ────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
log = logging.getLogger(__name__)

# ── Environment & Keys ───────────────────────────────────────────────────────
TELEGRAM_TOKEN: str = os.environ.get("TELEGRAM_BOT_TOKEN", "YOUR_TELEGRAM_TOKEN")
keys_env = os.environ.get("GEMINI_API_KEYS", os.environ.get("GEMINI_API_KEY", ""))
GEMINI_KEYS = [k.strip() for k in keys_env.split(",") if k.strip()]

if not GEMINI_KEYS:
    raise ValueError("No Gemini API keys found. Please set GEMINI_API_KEYS.")

_key_cycle = itertools.cycle(GEMINI_KEYS)
_KEY_LOCK = threading.Lock()

def get_next_key() -> str:
    with _KEY_LOCK:
        return next(_key_cycle)

PROXY_URL: Optional[str] = None
bot = telebot.TeleBot(TELEGRAM_TOKEN, parse_mode="Markdown")

# ── Constants ────────────────────────────────────────────────────────────────
OCR_MODEL       = "gemini-3.1-flash-lite"
MAX_RETRIES     = 6
BACKOFF_BASE    = 1.5
BACKOFF_CAP     = 64.0
SESSION_TTL     = 3600

PHONE_RE = re.compile(r"^0[79]\d{8}$")
FILES_UPLOAD_URL = "https://generativelanguage.googleapis.com/upload/v1beta/files?uploadType=multipart"

# ── State Management ─────────────────────────────────────────────────────────
class State:
    IDLE = "IDLE"
    PROCESSING = "PROCESSING"
    AWAITING_CONFIRM = "AWAITING_CONFIRM"

class PhotoRecord:
    def __init__(self, photo_id: int, start_ticket: str, data: dict):
        self.photo_id = photo_id
        self.start_ticket = start_ticket
        self.data = data

class Session:
    def __init__(self):
        self.state = State.IDLE
        self.photos: List[PhotoRecord] = []
        self.last_active = time.time()

    def reset(self):
        self.state = State.IDLE
        self.photos.clear()
        self.last_active = time.time()

_sessions: Dict[int, Session] = {}

def get_session(chat_id: int) -> Session:
    if chat_id not in _sessions:
        _sessions[chat_id] = Session()
    _sessions[chat_id].last_active = time.time()
    return _sessions[chat_id]

def _next_photo_id(sess: Session) -> int:
    return max([p.photo_id for p in sess.photos] + [0]) + 1

# ── Prompt & Parser ──────────────────────────────────────────────────────────
_PROMPT = """
You are a specialist OCR assistant for Ethiopian handwritten lottery logbooks.
You will receive multiple photos of logbook pages.

PAGE LAYOUT — each page has one or more vertical column-pairs:
  LEFT  : sequential ticket numbers (handwritten integers, ascending).
  RIGHT : customer name in Amharic (IGNORE) and a 10-digit Ethiopian mobile number starting with 09 or 07.

SPANNING RULE:
  A phone number written large may visually span several ticket rows.
  Assign that SAME number to EVERY ticket it covers. Never mark a covered row "Empty".

VALIDATION:
  • Phone must be exactly 10 digits, starting with 07 or 09.
  • If unreadable or absent → "Empty". Never guess or invent digits.

OUTPUT FORMAT:
Return ONLY a valid JSON object. No markdown. No prose.
Group the output by the image labels provided in the prompt (e.g., image_1, image_2).
{"image_1": [{"ticket": "1", "phone": "0916039018"}], "image_2": [{"ticket": "2", "phone": "Empty"}]}
"""

def _parse_ocr(raw: str) -> dict[str, dict[str, str]]:
    raw = re.sub(r"^\s*```[a-zA-Z]*\s*", "", raw.strip())
    raw = re.sub(r"\s*```\s*$", "", raw).strip()
    
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        raise ValueError(f"No JSON object found: {raw[:200]}")
    
    parsed = json.loads(match.group(0))
    result = {}
    
    for img_key, tickets_list in parsed.items():
        if not isinstance(tickets_list, list): continue
        img_result = {}
        for item in tickets_list:
            if not isinstance(item, dict): continue
            ticket = str(item.get("ticket", "")).strip()
            phone  = str(item.get("phone",  "Empty")).strip()
            if not ticket: continue
            if phone != "Empty" and not PHONE_RE.match(phone): phone = "Empty"
            img_result[ticket] = phone
        result[img_key] = img_result
    return result

# ── Files API Helpers ────────────────────────────────────────────────────────
def _upload_image(image_bytes: bytes, api_key: str, mime_type: str = "image/jpeg") -> str:
    proxies = {"https": PROXY_URL, "http": PROXY_URL} if PROXY_URL else None
    metadata_part = json.dumps({"file": {"display_name": "logbook_page"}})
    boundary = "-----lottery_bot_boundary"
    
    body = (
        f"--{boundary}\r\n"
        f"Content-Type: application/json; charset=utf-8\r\n\r\n"
        f"{metadata_part}\r\n"
        f"--{boundary}\r\n"
        f"Content-Type: {mime_type}\r\n\r\n"
    ).encode() + image_bytes + f"\r\n--{boundary}--".encode()

    headers = {
        "X-Goog-Api-Key": api_key,
        "Content-Type": f"multipart/related; boundary={boundary}",
        "Content-Length": str(len(body)),
    }

    resp = requests.post(FILES_UPLOAD_URL, headers=headers, data=body, proxies=proxies, timeout=60)
    resp.raise_for_status()
    return resp.json()["file"]["uri"]

def _delete_file(file_uri: str, api_key: str) -> None:
    try:
        proxies = {"https": PROXY_URL, "http": PROXY_URL} if PROXY_URL else None
        name = file_uri.split("/v1beta/")[-1] if "/v1beta/" in file_uri else file_uri
        url = f"https://generativelanguage.googleapis.com/v1beta/{name}"
        requests.delete(url, headers={"X-Goog-Api-Key": api_key}, proxies=proxies, timeout=10)
    except Exception:
        pass

# ── Core AI Extractor ────────────────────────────────────────────────────────
def extract_tickets(images_bytes: list[bytes]) -> dict[str, dict[str, str]]:
    uploaded_uris = []
    current_key = get_next_key()
    
    try:
        for image_bytes in images_bytes:
            mime_type = "image/jpeg"
            if image_bytes.startswith(b"\x89PNG"): mime_type = "image/png"
            elif image_bytes.startswith(b"\xff\xd8\xff"): mime_type = "image/jpeg"
            elif image_bytes.startswith(b"RIFF") and b"WEBP" in image_bytes[:12]: mime_type = "image/webp"
            elif image_bytes.startswith(b"GIF8"): mime_type = "image/gif"
            
            for attempt in range(3):
                try:
                    file_uri = _upload_image(image_bytes, current_key, mime_type=mime_type)
                    uploaded_uris.append((file_uri, mime_type))
                    break
                except Exception as e:
                    if attempt == 2: raise
                    time.sleep(2 ** attempt)

        with _KEY_LOCK:
            genai.configure(api_key=current_key)
            
            for file_uri, _ in uploaded_uris:
                for _ in range(10):
                    file_info = genai.get_file(file_uri.split("/")[-1])
                    if file_info.state.name == "ACTIVE": break
                    time.sleep(1)

            model = genai.GenerativeModel(OCR_MODEL)
            prompt_parts = [_PROMPT]
            for idx, (file_uri, mime_type) in enumerate(uploaded_uris, start=1):
                prompt_parts.append(f"image_{idx}:")
                prompt_parts.append({"file_data": {"mime_type": mime_type, "file_uri": file_uri}})

            sleep_time = BACKOFF_BASE
            last_exc = None

            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    response = model.generate_content(
                        prompt_parts,
                        generation_config={"temperature": 0.1, "max_output_tokens": 8192},
                    )
                    return _parse_ocr(response.text)
                except (json.JSONDecodeError, ValueError) as e:
                    last_exc = e
                    if attempt >= 2: raise RuntimeError(f"Bad JSON twice: {e}") from e
                    continue
                except Exception as e:
                    last_exc = e
                    err_str = str(e)
                    if "429" in err_str or "quota" in err_str.lower() or "503" in err_str:
                        m = re.search(r"retry[_ ]delay.*?(\d+\.?\d*)", err_str, re.I)
                        actual_sleep = min(float(m.group(1)) + 1.0 if m else sleep_time, BACKOFF_CAP)
                        time.sleep(actual_sleep)
                        sleep_time = min(sleep_time * 2, BACKOFF_CAP)
                        continue
                    raise
            raise RuntimeError(f"OCR failed after {MAX_RETRIES} attempts. Last: {last_exc}") from last_exc
    finally:
        for file_uri, _ in uploaded_uris:
            _delete_file(file_uri, current_key)

# ── File Export Builders ─────────────────────────────────────────────────────
def _tsort(t: str):
    return int(t) if t.isdigit() else t

def _merged(photos: list[PhotoRecord]) -> list[tuple[str, str]]:
    m: dict[str, str] = {}
    for ph in photos:
        m.update(ph.data)
    return sorted(m.items(), key=lambda kv: _tsort(kv[0]))

def build_excel(photos: list[PhotoRecord]) -> bytes:
    rows = _merged(photos)
    df = pd.DataFrame([{"No.": i, "Phone Number": ("" if p == "Empty" else p)} for i, (_, p) in enumerate(rows, 1)])
    buf = io.BytesIO()
    df.to_excel(buf, index=False, sheet_name="Lottery Sales")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = bdr
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center; cell.border = bdr

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def build_csv(photos: list[PhotoRecord]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["No.", "Phone Number"])
    for i, (_, p) in enumerate(_merged(photos), 1):
        writer.writerow([i, "" if p == "Empty" else p])
    return buf.getvalue().encode("utf-8-sig")

def build_txt(photos: list[PhotoRecord]) -> bytes:
    lines = [p for _, p in _merged(photos) if p != "Empty"]
    block = "\n\n--- comma-separated ---\n" + ", ".join(lines)
    return ("\n".join(lines) + block).encode("utf-8")

# ── Batch Processor ──────────────────────────────────────────────────────────
def batch_summary(photos: List[PhotoRecord]) -> str:
    total_tickets = sum(len(p.data) for p in photos)
    return f"📊 *Batch Summary*\nPhotos processed: {len(photos)}\nTotal tickets: {total_tickets}"

def export_kb(add_more: bool = True) -> types.InlineKeyboardMarkup:
    kb = types.InlineKeyboardMarkup(row_width=1)
    if add_more:
        kb.add(types.InlineKeyboardButton("📷 Add More Photos", callback_data="add_more"))
    kb.add(types.InlineKeyboardButton("📊 Export Excel (.xlsx)", callback_data="export_xlsx"))
    kb.add(types.InlineKeyboardButton("📄 Export CSV (.csv)", callback_data="export_csv"))
    kb.add(types.InlineKeyboardButton("📝 Export Text (.txt)", callback_data="export_txt"))
    kb.add(types.InlineKeyboardButton("❌ Cancel & Discard", callback_data="cancel"))
    return kb

def _process_batch(chat_id: int, file_ids: list[str]) -> None:
    sess = get_session(chat_id)
    total = len(file_ids)

    try:
        progress = bot.send_message(chat_id, f"⚙️ Downloading *{total}* photo(s) to process as a single batch…")

        image_bytes_list = []
        for file_id in file_ids:
            info = bot.get_file(file_id)
            image_bytes_list.append(bot.download_file(info.file_path))

        bot.edit_message_text(f"⚙️ Running AI OCR on all *{total}* photos at once...", chat_id, progress.message_id)

        try:
            batch_data = extract_tickets(image_bytes_list)
        except Exception as exc:
            log.error("OCR failed batch for chat=%d: %s", chat_id, exc)
            if "429" in str(exc) or "quota" in str(exc).lower(): 
                msg = "⚠️ The AI service is busy. Please try again."
            else: 
                msg = "❌ Could not process this batch. Try clearer images."
            bot.send_message(chat_id, msg)
            return

        for idx in range(1, total + 1):
            data = batch_data.get(f"image_{idx}")
            if not data and len(batch_data) >= idx:
                data = list(batch_data.values())[idx - 1]

            if not data:
                continue

            try: start = str(min((int(t) for t in data if t.isdigit()), default=0))
            except Exception: start = next(iter(data), "?")

            sess.photos.append(PhotoRecord(_next_photo_id(sess), start, data))

        try: bot.delete_message(chat_id, progress.message_id)
        except Exception: pass

        sess.state = State.AWAITING_CONFIRM if sess.photos else State.IDLE

        if sess.photos: 
            bot.send_message(chat_id, batch_summary(sess.photos), reply_markup=export_kb())
        else: 
            bot.send_message(chat_id, "⚠️ No data could be extracted.")

    except Exception as exc:
        log.error("Crash in _process_batch: %s", exc)
        try: bot.send_message(chat_id, "❌ An unexpected error occurred.")
        except Exception: pass
    finally:
        if sess.state == State.PROCESSING: sess.state = State.IDLE

# ── Handlers ─────────────────────────────────────────────────────────────────
@bot.callback_query_handler(func=lambda c: True)
def handle_callback(call: types.CallbackQuery) -> None:
    chat_id = call.message.chat.id
    sess = get_session(chat_id)
    bot.answer_callback_query(call.id)
    action = call.data

    if action == "add_more":
        if sess.state != State.AWAITING_CONFIRM: return
        sess.state = State.IDLE
        try: bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)
        except Exception: pass
        bot.send_message(chat_id, "📷 Ready — send your next photo(s).")
        return

    if action in ("export_xlsx", "export_csv", "export_txt"):
        if not sess.photos or sess.state != State.AWAITING_CONFIRM:
            bot.send_message(chat_id, "⚠️ No data to export.")
            return
        building = bot.send_message(chat_id, "⏳ Building your file…")
        try:
            if action == "export_xlsx":
                raw = build_excel(sess.photos)
                bot.send_document(chat_id, io.BytesIO(raw), visible_file_name="lottery_tickets.xlsx", caption="📊 Excel export ready.")
            elif action == "export_csv":
                raw = build_csv(sess.photos)
                bot.send_document(chat_id, io.BytesIO(raw), visible_file_name="lottery_tickets.csv", caption="📄 CSV export ready.")
            else:
                raw = build_txt(sess.photos)
                bot.send_document(chat_id, io.BytesIO(raw), visible_file_name="lottery_tickets.txt", caption="📝 Text export ready.")
            
            try: bot.delete_message(chat_id, building.message_id)
            except Exception: pass
            try: bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)
            except Exception: pass
            sess.reset()
        except Exception as exc:
            log.error("Export failed: %s", exc)
            bot.edit_message_text("❌ Export failed.", chat_id, building.message_id)
        return

    if action == "cancel":
        sess.reset()
        try: bot.edit_message_text("❌ Batch cancelled.", chat_id, call.message.message_id)
        except Exception: bot.send_message(chat_id, "❌ Batch cancelled.")

@bot.message_handler(content_types=['photo'])
def handle_photos(msg: types.Message):
    sess = get_session(msg.chat.id)
    if sess.state == State.PROCESSING:
        bot.reply_to(msg, "⏳ Please wait, processing a batch...")
        return
        
    sess.state = State.PROCESSING
    file_id = msg.photo[-1].file_id
    Thread(target=_process_batch, args=(msg.chat.id, [file_id]), daemon=True).start()

@bot.message_handler(commands=["start", "help"])
def handle_start(msg: types.Message) -> None:
    bot.send_message(msg.chat.id, "👋 Send me photos of your logbook to extract tickets.")

@bot.message_handler(commands=["cancel"])
def handle_cancel(msg: types.Message) -> None:
    get_session(msg.chat.id).reset()
    bot.send_message(msg.chat.id, "❌ Session cancelled.")

@bot.message_handler(commands=["status"])
def handle_status(msg: types.Message) -> None:
    sess = get_session(msg.chat.id)
    if not sess.photos:
        bot.send_message(msg.chat.id, "ℹ️ No active batch.")
        return
    bot.send_message(msg.chat.id, batch_summary(sess.photos), reply_markup=export_kb(add_more=sess.state == State.AWAITING_CONFIRM))

@bot.message_handler(func=lambda m: get_session(m.chat.id).state == State.IDLE, content_types=["text"])
def handle_idle(msg: types.Message) -> None:
    bot.send_message(msg.chat.id, "📸 Send a photo of your logbook page to get started.")

# ── Render Dummy Web Server ──────────────────────────────────────────────────
import http.server
import socketserver

def run_web_server():
    port = int(os.environ.get("PORT", 10000))
    Handler = http.server.SimpleHTTPRequestHandler
    try:
        with socketserver.TCPServer(("", port), Handler) as httpd:
            httpd.serve_forever()
    except Exception: pass

web_thread = Thread(target=run_web_server, daemon=True)
web_thread.start()

# ── Polling Entry Point ──────────────────────────────────────────────────────
if __name__ == "__main__":
    log.info("Starting Lottery Logbook Bot...")
    bot.remove_webhook() # Clears conflicting webhooks
    while True:
        try:
            bot.polling(none_stop=True, timeout=60)
        except Exception as e:
            log.error("Bot polling failed: %s. Restarting in 5s...", e)
            time.sleep(5)
